from openpyxl import load_workbook
wb = load_workbook('SLR(1)2.xlsx')
sheet = wb[wb.sheetnames[-1]]
rows = sheet.rows
head = next(rows)[:22]
head = [i.value for i in head]
action_head = head[1:17]
goto_head = head[17:22]
print(action_head,goto_head)
print('-'*22)
action=[]
goto=[]
for i in range(36):
    line =next(rows)[:22]
    line = [i.value for i in line]
    # print(line)
    action .append(line[1:17])
    goto.append(line[17:22])
# print(action,goto)
print(f"const std::string action_table[][{len(action[0])}]={{")
for a in action:
    print("    {",end='')
    for i in a:
        print(f'"{i if i else "x"}",',end='')
    print("},")
print("};")

print(f"const int goto_table[][{len(goto[0])}]={{")
for a in goto:
    print("    {",end='')
    for i in a:
        print(f'{i if i else -1},',end='')
    print("},")
print("};")
